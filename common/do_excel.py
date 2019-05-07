# -*- coding:utf-8 -*-
# @Time  : 2019/5/5 2:58
# @Author: xiaoxiao
# @File  : do_excel.py

from openpyxl import load_workbook

class Case:
    case_id=None
    title=None
    url=None
    data=None
    expected=None
    check_sql=None


class DoExcel:
    def __init__(self,filename,sheetname):
        self.filename=filename
        self.workbook=load_workbook(filename)
        self.sheetname=self.workbook[sheetname]

    def read(self):
        row = self.sheetname.max_row
        cases = []
        for row in range(2, row + 1):
            case = Case()
            case.case_id = self.sheetname.cell(row, 1).value
            case.title = self.sheetname.cell(row, 2).value
            case.url = self.sheetname.cell(row, 3).value
            case.data = self.sheetname.cell(row, 4).value
            case.expected = self.sheetname.cell(row, 5).value
            case.check_sql = self.sheetname.cell(row, 8).value
            cases.append(case)
        return cases
    def write(self,row,column,str):
        self.sheetname.cell(row,column).value=str
    def close(self):
        self.workbook.save(self.filename)
        self.workbook.close()



